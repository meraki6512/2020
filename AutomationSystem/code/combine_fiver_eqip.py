# mysql
import pymysql

#exel
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

myhost='localhost'
myuser='root'
mypassword='0108'
mycharset='utf8'

exel_fiver_file="C:\\Users\\LEE CHAN HO\\PycharmProjects\\kt\\workfile\\200801_fiver_list.xlsx"
exel_fiver_sheet="200801_fiver_list"

exel_eqip_file="C:\\Users\\LEE CHAN HO\\PycharmProjects\\kt\\workfile\\200801_tr_equip.xlsx"
exel_eqip_sheet="Sheet1"
#0.1로그인 함수 -> db에 접속하는 함수
def login_mysql(login_host,login_user,login_password,login_charset):
    conn = pymysql.connect(host=login_host, user=login_user, password=login_password, charset=login_charset)
    return conn

def login_mysql_db(login_host,login_user,login_password,db_name,login_charset):
    conn = pymysql.connect(host=login_host, user=login_user, password=login_password, db=db_name ,charset=login_charset)
    return conn
#로그인과 동시에 특정 db에 연결하는 함수
#0.2 탐색 함수
def search_db(db_name):
    #인자인 db_name의 이름이 데이터 베이스 내에 있는지 확인하는 함수
    conn = login_mysql(login_host=myhost, login_user=myuser, login_password=mypassword, login_charset=mycharset)
    mycursor = conn.cursor()

    sql_search = "SHOW DATABASES LIKE '%s'" % db_name
    mycursor.execute(sql_search)
    ret = mycursor.fetchall()

    conn.commit()
    conn.close()
    #ret에는 데이터 베이스에서 db_name으로 검색한 결과가 담겨있다.
    if len(ret)==1:
        print('1')
        return 1
    elif len(ret)==0:
        print('0')
        return 0
    else:
        return -1
    #만약 DB_name이 없는 데이터 베이스라면 빈 ret 튜플이 나온다. 즉 비었다면 없는 것

#table은 변수명으로 받을 수 있다.
def search_table(db_name,table_name):
    conn = login_mysql(login_host=myhost, login_user=myuser, login_password=mypassword, login_charset=mycharset)
    mycursor = conn.cursor()
    sql_search_table="SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA = '%s' AND TABLE_NAME LIKE '%s'" % (db_name,table_name)
    mycursor.execute(sql_search_table)
    ret = mycursor.fetchall()
    conn.commit()
    conn.close()
    if len(ret)==1:
        print('1')
        return 1
    elif len(ret)==0:
        print('0')
        return 0
    else:
        return -1
#데이터 베이스 검색 함수랑 같은 원리

#0.3 make db & table function
def make_db(db_name):
    conn = login_mysql(login_host = myhost,login_user=myuser,login_password=mypassword,login_charset=mycharset)
    mycursor = conn.cursor()
    # 스케마 찾기부터 해야함.
    ret = search_db(db_name)
    if ret == 0:
        # ret == 0 이라는 건 kt_db라는 이름의 데이터베이스가 없다는 뜻이므로 아래 코드와 같이 생성해 준다.
        #sql ='CREATE SCHEMA %s DEFAULT CHARSET utf8 COLLATE utf8_bin'%db_name <- 이런식으로 짜는 이유는 ''없애기 위해서
        sql ='CREATE SCHEMA %s DEFAULT CHARSET utf8 COLLATE utf8_bin'%db_name
        mycursor.execute(sql)
        conn.commit()
        conn.close()
        print("success make db")
        return 0;

    elif ret == 1:
        print("already")
        conn.commit()
        conn.close()
        print("please rename DB_name")
        return 1

    else:
        print("find error")
        conn.commit()
        conn.close()
        return -1

def make_table(db_name,table_name):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name= db_name,login_charset=mycharset)
    mycursor = conn.cursor()
    #테이블 찾기
    ret_table = search_table(db_name=db_name, table_name=table_name)
    if ret_table==0:
        sql = "CREATE TABLE `%s`(`id` INT NOT NULL AUTO_INCREMENT, PRIMARY KEY (`id`))"%table_name
        mycursor.execute(sql)
        conn.commit()
        conn.close()
        print("make_table")
        return 0
    elif ret_table==1:
        print("already")
        conn.commit()
        conn.close()
        print("please rename Table_name")
        return 1
        #테이블 내용 제거 필요!!
    else:
        print("find error")
        conn.commit()
        conn.close()
        return -1

def rename_table_plus(db_name,table_name):
    conn = login_mysql(login_host=myhost, login_user=myuser, login_password=mypassword, login_charset=mycharset)
    mycursor = conn.cursor()


#0.4 Column
def add_int_col(db_name,table_name,col_name):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    sql="alter table `%s` add(`%s` int null)" % (table_name,col_name)
    mycursor.execute(sql)

def add_char_col(db_name,table_name,col_name):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    sql="alter table `%s` add(`%s` varchar (45) null)"%(table_name,col_name)
    mycursor.execute(sql)

def add_char_fiver_col(db_name,table_name,col_name):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    sql = "alter table `%s` add(`%s` varchar (300) null)" % (table_name, col_name)
    mycursor.execute(sql)

def search_col(db_name,table_name,col_name):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    #sql="SELECT `%s` from `%s`"%(col_name,table_name)
    sql="SHOW columns FROM `%s` like '%s'" % (table_name,col_name)
    mycursor.execute(sql)
    col=mycursor.fetchall()

    #print(col)

    if len(col)==1:
        print("This column exist.")
        return 1
    elif len(col)==0:
        print("This column does not exist.")
        return 0

#0.4 db 제거함수
def delete_db(db_name):
    conn = login_mysql(login_host=myhost, login_user=myuser, login_password=mypassword, login_charset=mycharset)
    mycursor = conn.cursor()
    ret=search_db(db_name)
    if ret==1:
        sql="DROP DATABASES `%s`" % db_name
        mycursor.execute(sql)
        print("delete_db")
    elif ret==0:
        print("This \'%s\' does not exist."%db_name)

def delete_table(db_name,table_name):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    ret=search_table(db_name,table_name)
    if ret ==1:
        sql="DROP TABLE `%s`"%table_name
        mycursor.execute(sql)
        print("delete_table \'%s\' " % table_name)
    elif ret == 0:
        print("This \'%s\' does not exist."%table_name)

def show_database():
    conn = login_mysql(login_host=myhost, login_user=myuser, login_password=mypassword, login_charset=mycharset)
    mycursor = conn.cursor()
    sql="SHOW DATABASES"
    mycursor.execute(sql)
    data=mycursor.fetchall()
    for i in data:
        print(i)

def show_tables(db_name):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()

    sql="SHOW TABLES"
    mycursor.execute(sql)
    data=mycursor.fetchall()
    for i in data:
        print(i)
    return data

#0.5 fiver eqip col
def make_fiver_col(db_name, table_name, exel_name, exel_sheet_name):
    wx = openpyxl.load_workbook(exel_name)
    sheet = wx.get_sheet_by_name(exel_sheet_name)

    maxrow = sheet.max_row
    maxcol = sheet.max_column

    Kategorie = []
    kate_index = []
    count_index = 0

    for i in sheet.iter_rows(min_row=8, max_row=8, min_col=1, max_col=maxcol):
        for cell in i:
            kate_index.append(count_index)
            count_index += 1
            Kategorie.append(cell.value)

    search_col_ret = search_col(db_name, table_name, Kategorie[0])
    if search_col_ret==1:
        print("already fiver_coll exist")
        return 1

    for k in kate_index:
        if k==19:
            add_char_fiver_col(db_name,table_name,Kategorie[k])
            continue
        add_char_col(db_name,table_name,Kategorie[k])
    print("seucces make fiver col")

def make_eqip_col(db_name,table_name,exel_name,exel_sheet_name):
    wx = openpyxl.load_workbook(exel_name)
    sheet = wx.get_sheet_by_name(exel_sheet_name)

    maxrow = sheet.max_row
    maxcol = sheet.max_column

    format_count = int((maxcol - 1) / 3)

    eqip_format = ['국사', 'IP-Address', '장비명']
    format_eqip_index = []
    eqip_index_version = []

    search_col_ret = search_col(db_name, table_name, "캐리어")
    if search_col_ret == 1:
        print("already eqip_coll exist")
        return 1

    for i in range(1, format_count + 1):
        tmp = '(%d)' % i
        format_eqip_index.append(tmp)

    for i in range(0, format_count):
        for k in range(len(eqip_format)):
            eqip_index_version.append(eqip_format[k] + format_eqip_index[i])

    add_char_col(db_name, table_name, "캐리어")
    for k in range(len(eqip_index_version)):
        add_char_col(db_name, table_name, eqip_index_version[k])

    print("seucces make_eqip_col")

def make_display_col(db_name,table_name,exel_name,exel_sheet_name):
    wx = openpyxl.load_workbook(exel_name)
    sheet = wx.get_sheet_by_name(exel_sheet_name)

    maxrow = sheet.max_row
    maxcol = sheet.max_column

    format_count = int((maxcol-1)/3)

    eqip_format = ['국사', '장비명', '공사 전 일시', '공사 전 레벨', '공사 후 일시', '공사 후 레벨', '광레벨 편차', '조치 사항']
    format_eqip_index = []
    eqip_index_version = []

    search_col_ret = search_col(db_name, table_name, "광 케이블명")
    if search_col_ret == 1:
        print("already all_display_coll exist")
        return 1

    for i in range(1, format_count + 1):
        tmp = '(%d)' % i
        format_eqip_index.append(tmp)

    for i in range(0, format_count):
        for k in range(len(eqip_format)):
            eqip_index_version.append(eqip_format[k] + format_eqip_index[i])
    add_char_col(db_name, table_name, "광 케이블명")
    add_char_col(db_name, table_name, "코어번호")
    add_char_col(db_name,table_name,"캐리어/전송로")
    for k in range(len(eqip_index_version)):
        add_char_col(db_name,table_name,eqip_index_version[k])

    print("seucces make_all_db_col")

def inquiry_col(db_name,table_name):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    sql="SELECT column_name FROM information_schema.columns WHERE table_schema = '%s' AND table_name = '%s'"%(db_name,table_name)
    mycursor.execute(sql)
    col_list = mycursor.fetchall()
    return col_list

#0.6e 전용 데이터 읽기
def read_eqip_data(exel_eqip_file, exel_eqip_sheet):
    # 접속
    wx = openpyxl.load_workbook(exel_eqip_file)
    sheet = wx.get_sheet_by_name(exel_eqip_sheet)
    # 기본 데이터
    maxrow = sheet.max_row  # 7=1(카테고리)+6(값)
    maxcol = sheet.max_column  # 16

    # row 개수 구하기
    data = []
    count_index = []
    count = 0
    # 차원추가 -> 1 row당 1차원
    for k in sheet.iter_rows(min_row=2, max_row=maxrow, min_col=1, max_col=1):
        for cell in k:
            count_index.append(count)
            data.append([])
        count += 1

    # 각 차원에 값넣기
    s_count = 0
    for i in sheet.iter_rows(min_row=2, max_row=maxrow, min_col=1, max_col=maxcol):
        for cell in i:
            data[s_count].append(cell.value)
        s_count += 1

    del_index = range(3, len(data[0]), 3)

    # 데이터 빵꾸 채우기
    select_index = [x for x in range(0, len(data[0])) if x not in del_index]
    for i in range(1, len(data) + 1, 2):
        for j in select_index:
            data[i][j] = data[i - 1][j]

    return data

def insert_reset(db_name,table_name,exel_data,exel_col_list):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    # row제작을 위해서 만드는 첫 column => eiqp파일에서는 캐리어가 될것
    tmp='\0'
    for k in exel_col_list:
        tmp = list(k)
        if tmp[0] == 'id':
            continue
        break
    print(tmp[0])

    for i in range(len(exel_data)):
        sql = "insert into %s.`%s` (`%s`) values ('%s')" % (db_name, table_name, tmp[0], "NULL")

        mycursor.execute(sql)

    conn.commit()

    print('success insert_%s_reset'%table_name)

def update_reset(db_name, table_name, exel_data, exel_col_list):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    # row제작을 위해서 만드는 첫 column => eiqp파일에서는 캐리어가 될것
    tmp = '\0'
    for k in exel_col_list:
        tmp = list(k)
        if tmp[0] == 'id':
            continue
        break
    print(tmp[0])

    i = 1
    for k in range(len(exel_data)):
        sql = "update %s.`%s` set `%s` = %s where id = %d" % (db_name, table_name, tmp[0], 'NULL', i)
        i += 1
        mycursor.execute(sql)
    # 뭐 넣고나서는 무조건 conn.commit 써줘야 한다.
    conn.commit()

    print('success update_%s_reset' % table_name)

def inquiry_id_values(db_name,table_name):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()

    sql='select id from %s.`%s`'%(db_name,table_name)
    ret=mycursor.execute(sql)
    print('inquiry_id_values',ret)
    return ret

def eqip_update(db_name,table_name,eqip_data,eqip_col_list):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()

    ret=inquiry_id_values(db_name,table_name)
    if ret==0:
        insert_reset(db_name,table_name,eqip_data,eqip_col_list)
    elif ret>0:
        update_reset(db_name,table_name,eqip_data,eqip_col_list)

    for num,i in enumerate(eqip_col_list,0):
        tmp=list(i)
        if num==0:
            continue
        #print(tmp[0])
        #print(num)
        for j in range(len(eqip_data)):
                #print(tmp[0],eqip_data[j][num-1],'j=',j,'num=',num)
                sql="update %s.`%s` set `%s` = '%s' where id = %d" % (db_name,table_name,tmp[0], eqip_data[j][num-1], j+1)
                mycursor.execute(sql)

        conn.commit()

#전제: table name 검사를 통해서 중복된 이름이 없는 exel 이름이라는 전제
def eqip_exel_into_db(db_name,table_name,exel_name,exel_sheet):
    #1. 로그인
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name, login_charset=mycharset)
    mycursor = conn.cursor()
    print("success login")
    #2. 테이블 검색
    search_table_ret=search_table(db_name,table_name)
    if search_table_ret==0:
        make_table(db_name,table_name)
        print('%s에 해당하는 테이블을 만들었습니다.'%table_name)
    elif search_table_ret>0:
        print('테이블 이름이 %s인 테이블은 파일 관리에 `치명적인 오류`를 발생시킬 수 있습니다.'%table_name)
        print('파일 재확인하고 저장할 테이블 이름을 다시 입력하세요.')
        return 0
    print("success %s search"%table_name)
    #3. 데이터 읽기
    eqip_data=read_eqip_data(exel_name,exel_sheet)
    print("success %s_data_read"%table_name)
    #4. 컬럼 만들기
    make_eqip_col(db_name,table_name,exel_name,exel_sheet)
    print("success %s_db_column"%table_name)
    #4.5 컬럼 목록 불러오기
    eqip_col_list=inquiry_col(db_name, table_name)
    print("success %s_inquiry_col"%table_name)
    #5. 데이터 reset
    insert_reset(db_name, table_name, eqip_data, eqip_col_list)
    print("success %s_data_reset" % table_name)
    #6. 데이터 넣기
    eqip_update(db_name, table_name, eqip_data, eqip_col_list)
    print('성공적으로 %s데이터베이스(테이블)을 만들었습니다.' % table_name)
    print('--------------------------------------------------------------')

#0.7 fiver 전용 데이터 읽기
    #row=1 ~ row=7 기타 정보
    #row=8 카테고리
    #row=9~ max_row 데이터
def read_fiver_data(exel_fiver_file,exel_fiver_sheet):
    # 접속
    wx = openpyxl.load_workbook(exel_fiver_file)
    sheet = wx.get_sheet_by_name(exel_fiver_sheet)
    # 기본 데이터
    maxrow = sheet.max_row  #80
    maxcol = sheet.max_column  #38

    #다가져올거임 모든 row값에 대한 차원이 있어야함
    data=[]
    for i in sheet.iter_rows(min_row=9,max_row=maxrow,min_col=1,max_col=1):
            data.append([])

    count=0
    for i in sheet.iter_rows(min_row=9,max_row=maxrow,min_col=1,max_col=maxcol):
        for cell in i:
            data[count].append(cell.value)
        count+=1

    return data

def fiver_update(db_name,table_name,fiver_data,fiver_col_list):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()

    ret = inquiry_id_values(db_name, table_name)
    if ret == 0:
        insert_reset(db_name, table_name, fiver_data, fiver_col_list)
    elif ret > 0:
        update_reset(db_name, table_name, fiver_data, fiver_col_list)

    for num, i in enumerate(fiver_col_list, 0):
        tmp = list(i)
        if num == 0:
            continue
        # print(tmp[0])
        # print(num)
        if num-1==19:
            continue
        for j in range(len(fiver_data)):
            # print(tmp[0],eqip_data[j][num-1],'j=',j,'num=',num)
            sql = "update %s.`%s` set `%s` = '%s' where id = %d" % (db_name, table_name, tmp[0], fiver_data[j][num - 1], j + 1)
            mycursor.execute(sql)

        conn.commit()

def fiver_exel_into_db(db_name,table_name,exel_name,exel_sheet):
    # 1. 로그인
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    print("success login")
    # 2. 테이블 검색
    search_table_ret = search_table(db_name, table_name)
    if search_table_ret == 0:
        make_table(db_name, table_name)
        print('%s에 해당하는 테이블을 만들었습니다.' % table_name)
    elif search_table_ret > 0:
        print('테이블 이름이 %s인 테이블은 파일 관리에 `치명적인 오류`를 발생시킬 수 있습니다.' % table_name)
        print('파일 재확인하고 저장할 테이블 이름을 다시 입력하세요.')
        return 0
    print("success %s search" % table_name)
    # 3. 데이터 읽기
    fiver_data = read_fiver_data(exel_name, exel_sheet)
    print("success %s_data_read" % table_name)
    # 4. 컬럼 만들기
    make_fiver_col(db_name, table_name, exel_name, exel_sheet)
    print("success %s_db_column" % table_name)
    #4.5 컬럼 목록 불러오기
    fiver_col_list=inquiry_col(db_name, table_name)
    print("success %s_inquiry_col"%table_name)
    # 5. 데이터 reset
    insert_reset(db_name, table_name, fiver_data, fiver_col_list)
    print("success %s_data_reset" % table_name)
    # 6. 데이터 넣기
    fiver_update(db_name, table_name, fiver_data, fiver_col_list)
    print('성공적으로 %s데이터베이스(테이블)을 만들었습니다.'%table_name)
    print('--------------------------------------------------------------')

#0.8 combine fiver and equip
def make_servertable_base(db_name,combine_table,server_table):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    sql='create table `%s`(select * from `%s` where mod(`코어번호`,2)= mod(`id`,2))'%(server_table,combine_table)
    mycursor.execute(sql)
    conn.commit()
    print('success servertable_base')

def make_combine_server(db_name,fiver_table,eqip_table,combine_table,server_table):
    conn = login_mysql_db(login_host=myhost, login_user=myuser, login_password=mypassword, db_name=db_name,
                          login_charset=mycharset)
    mycursor = conn.cursor()
    sql='create table `%s`(select * from (select `케이블명`, `코어번호`, `캐리어/회선명` from `%s`)a join `%s` on %s.`캐리어` = a.`캐리어/회선명`)'\
        %(combine_table,fiver_table,eqip_table,eqip_table)
    print(sql)
    mycursor.execute(sql)
    conn.commit()

    print('success com table')
    make_servertable_base(db_name,combine_table,server_table)


    print('success combine_table')


'''
delete_table("test","equip")
delete_table('test','fiver')
eqip_exel_into_db("test","eqip",exel_eqip_file,exel_eqip_sheet)
fiver_exel_into_db("test",'fiver',exel_fiver_file,exel_fiver_sheet)
'''


#delete_table('test','com(2)')
#combine_fiver_eqip('test','fiver','eqip','com(2)')
'''
delete_table('test','display')
make_table('test','display')
make_display_col('test','display',exel_eqip_file,exel_eqip_sheet)
'''
#make_combine_server('test','fiver','eqip','com','server')
#make_table('test','test_table')


